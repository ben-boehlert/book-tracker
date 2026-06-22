#!/usr/bin/env python3
"""Regenerate derived artifacts from books.json (the single source of truth).

Outputs:
  index.html                 read-only sortable/filterable viewer (data embedded)
  exports/Book_Inventory.xlsx  spreadsheet with a status column + summary
  exports/book_inventory.md    plain-text table

Usage: python3 build.py
"""
import json, os
from collections import Counter

ROOT = os.path.dirname(os.path.abspath(__file__))
books = json.load(open(os.path.join(ROOT, "books.json"), encoding="utf-8"))

# ---- validate ----
STATUSES = {"Finished", "Started", "Not started"}
ns = [b["n"] for b in books]
assert len(set(ns)) == len(ns), "duplicate ids"
for b in books:
    assert b["status"] in STATUSES, b
    assert isinstance(b["pages"], int) and b["pages"] > 0, b
    assert b["rel"] in (1, 2, 3, 4, 5), b
books.sort(key=lambda b: b["n"])
cnt = Counter(b["status"] for b in books)
pages = sum(b["pages"] for b in books)
print(f"{len(books)} books | {dict(cnt)} | {pages:,} pages")

# ---- index.html (read-only viewer) ----
VIEW = r'''<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Book Tracker</title>
<style>
  :root{--bg:#f7f7f5;--card:#fff;--ink:#23211e;--muted:#6b6760;--line:#e7e4df;
    --fin-bg:#c6efce;--fin-fg:#0a6b2b;--sta-bg:#bdd7ee;--sta-fg:#1f4e78;--not-bg:#ececec;--not-fg:#8a8a8a;}
  @media (prefers-color-scheme:dark){:root{--bg:#1c1b19;--card:#262521;--ink:#ece9e3;--muted:#a39e95;--line:#3a3833;
    --fin-bg:#1e4d2b;--fin-fg:#9be3ad;--sta-bg:#22405e;--sta-fg:#a9cdf0;--not-bg:#34322d;--not-fg:#9a958c;}}
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--ink);font:14px/1.45 -apple-system,BlinkMacSystemFont,"Segoe UI",Helvetica,Arial,sans-serif;}
  .wrap{max-width:1080px;margin:0 auto;padding:20px 16px 60px;}
  h1{font-size:22px;margin:0 0 2px;font-weight:650;}
  .sub{color:var(--muted);font-size:13px;margin-bottom:16px;}
  .bar{display:flex;flex-wrap:wrap;gap:8px;align-items:center;margin-bottom:14px;}
  input[type=search],select{font:inherit;color:var(--ink);background:var(--card);border:1px solid var(--line);border-radius:8px;padding:7px 10px;}
  input[type=search]{flex:1;min-width:160px;}
  button{font:inherit;cursor:pointer;border:1px solid var(--line);background:var(--card);color:var(--ink);border-radius:8px;padding:7px 12px;}
  .stats{display:flex;flex-wrap:wrap;gap:10px;margin-bottom:14px;}
  .stat{background:var(--card);border:1px solid var(--line);border-radius:10px;padding:9px 13px;min-width:96px;}
  .stat b{display:block;font-size:18px;font-weight:650;}
  .stat span{color:var(--muted);font-size:11.5px;text-transform:uppercase;letter-spacing:.04em;}
  .progress{height:8px;border-radius:5px;background:var(--not-bg);overflow:hidden;display:flex;margin:2px 0 16px;}
  .progress i{display:block;height:100%;}
  table{width:100%;border-collapse:collapse;background:var(--card);border:1px solid var(--line);border-radius:12px;overflow:hidden;}
  th,td{text-align:left;padding:9px 11px;border-bottom:1px solid var(--line);vertical-align:top;}
  th{font-size:12px;color:var(--muted);text-transform:uppercase;letter-spacing:.03em;cursor:pointer;user-select:none;white-space:nowrap;position:sticky;top:0;background:var(--card);}
  th.sorted::after{content:" \25B2";font-size:9px}
  th.sorted.desc::after{content:" \25BC"}
  tr:last-child td{border-bottom:none}
  td.num,th.num{text-align:right;font-variant-numeric:tabular-nums;}
  td.center,th.center{text-align:center}
  .title{font-weight:600}
  .auth{color:var(--muted);font-size:12.5px}
  .rel{display:inline-block;min-width:20px;text-align:center;font-weight:700;border-radius:6px;padding:1px 7px;font-size:12px;}
  .badge{display:inline-block;border-radius:999px;padding:3px 11px;font-size:12px;font-weight:600;white-space:nowrap;}
  .Finished{background:var(--fin-bg);color:var(--fin-fg);}
  .Started{background:var(--sta-bg);color:var(--sta-fg);}
  .ns{background:var(--not-bg);color:var(--not-fg);}
  .hint{color:var(--muted);font-size:12px;margin-top:10px;}
  .pill{font-size:11px;color:var(--muted);border:1px solid var(--line);border-radius:999px;padding:1px 8px;}
  .only-sm{display:none}
  @media (max-width:620px){.hide-sm{display:none}.only-sm{display:block}.wrap{padding:14px 10px 50px}}
</style>
</head>
<body>
<div class="wrap">
  <h1>Book Tracker</h1>
  <div class="sub">Read-only view. Source of truth is <code>books.json</code>; ask Claude to change a status, then it&rsquo;s committed to git.</div>
  <div class="stats" id="stats"></div>
  <div class="progress" id="prog"></div>
  <div class="bar">
    <input type="search" id="q" placeholder="Search title or author&hellip;">
    <select id="fStatus"><option value="">All statuses</option><option>Finished</option><option>Started</option><option>Not started</option></select>
    <select id="fCat"><option value="">All categories</option></select>
    <select id="fRel"><option value="">All relevance</option><option value="5">5 &middot; Core</option><option value="4">4 &middot; Strong</option><option value="3">3 &middot; Adjacent</option><option value="2">2 &middot; Tangential</option><option value="1">1 &middot; Personal</option></select>
    <button id="reset">Reset</button>
  </div>
  <table>
    <thead><tr>
      <th data-k="n" class="num">#</th><th data-k="title">Title</th><th data-k="author" class="hide-sm">Author</th>
      <th data-k="shelf" class="center hide-sm">Shelf</th><th data-k="cat" class="hide-sm">Category</th>
      <th data-k="rel" class="center">Rel</th><th data-k="pages" class="num">Pages</th><th data-k="status" class="center">Status</th>
    </tr></thead>
    <tbody id="body"></tbody>
  </table>
  <div class="hint">Relevance: <b>5</b> core &middot; <b>4</b> strongly related &middot; <b>3</b> adjacent &middot; <b>2</b> tangential &middot; <b>1</b> personal.</div>
</div>
<script type="application/json" id="books">__DATA__</script>
<script>
const RB={5:'#c6efce',4:'#d6e9c6',3:'#ffeb9c',2:'#fce4d6',1:'#e7e6e6'},RF={5:'#006100',4:'#375623',3:'#7f6000',2:'#843c0c',1:'#595959'};
const ORD={'Not started':0,'Started':1,'Finished':2};
const books=JSON.parse(document.getElementById('books').textContent);
let sortK='n',dir=1;
const cs=document.getElementById('fCat');
[...new Set(books.map(b=>b.cat))].sort().forEach(c=>{const o=document.createElement('option');o.value=o.textContent=c;cs.appendChild(o);});
const esc=s=>String(s).replace(/[&<>"]/g,c=>({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;'}[c]));
const cls=s=>s==='Not started'?'ns':s;
function filt(){const q=document.getElementById('q').value.toLowerCase().trim(),fs=fStatus.value,fc=fCat.value,fr=fRel.value;
  return books.filter(b=>(!fs||b.status===fs)&&(!fc||b.cat===fc)&&(!fr||String(b.rel)===fr)&&(!q||b.title.toLowerCase().includes(q)||b.author.toLowerCase().includes(q)));}
function render(){const rows=filt().slice().sort((a,b)=>{let x=a[sortK],y=b[sortK];if(sortK==='status'){x=ORD[x];y=ORD[y];}if(typeof x==='string'){x=x.toLowerCase();y=y.toLowerCase();}return(x<y?-1:x>y?1:0)*dir;});
  document.getElementById('body').innerHTML=rows.map(b=>'<tr><td class="num">'+b.n+'</td>'+
    '<td><span class="title">'+esc(b.title)+'</span><div class="auth only-sm">'+esc(b.author)+'</div></td>'+
    '<td class="auth hide-sm">'+esc(b.author)+'</td><td class="center hide-sm"><span class="pill">'+(b.shelf??'—')+'</span></td>'+
    '<td class="hide-sm">'+esc(b.cat)+'</td><td class="center"><span class="rel" style="background:'+RB[b.rel]+';color:'+RF[b.rel]+'">'+b.rel+'</span></td>'+
    '<td class="num">'+b.pages+'</td><td class="center"><span class="badge '+cls(b.status)+'">'+b.status+'</span></td></tr>').join('');
  document.querySelectorAll('th').forEach(th=>{th.classList.toggle('sorted',th.dataset.k===sortK);th.classList.toggle('desc',th.dataset.k===sortK&&dir<0);});}
function stats(){const t=books.length,pg=books.reduce((s,b)=>s+b.pages,0),by=s=>books.filter(b=>b.status===s);
  const f=by('Finished'),st=by('Started'),n=by('Not started'),pr=f.reduce((s,b)=>s+b.pages,0);
  document.getElementById('stats').innerHTML=[[t,'Books'],[f.length,'Finished'],[st.length,'Started'],[n.length,'Unread'],[pg.toLocaleString(),'Total pages'],[pr.toLocaleString(),'Pages read']].map(s=>'<div class="stat"><b>'+s[0]+'</b><span>'+s[1]+'</span></div>').join('');
  document.getElementById('prog').innerHTML='<i style="width:'+Math.round(f.length/t*100)+'%;background:var(--fin-fg)"></i><i style="width:'+Math.round(st.length/t*100)+'%;background:var(--sta-fg)"></i>';}
document.querySelectorAll('th').forEach(th=>th.onclick=()=>{const k=th.dataset.k;if(sortK===k)dir*=-1;else{sortK=k;dir=1;}render();});
['q','fStatus','fCat','fRel'].forEach(id=>document.getElementById(id).addEventListener('input',render));
document.getElementById('reset').onclick=()=>{q.value='';fStatus.value='';fCat.value='';fRel.value='';render();};
stats();render();
</script>
</body>
</html>'''
open(os.path.join(ROOT, "index.html"), "w", encoding="utf-8").write(
    VIEW.replace("__DATA__", json.dumps(books, ensure_ascii=False)))

# ---- exports/book_inventory.md ----
md = [f"# Book Inventory",
      f"",
      f"_{cnt['Finished']} finished &middot; {cnt['Started']} started &middot; {cnt['Not started']} not started &middot; {len(books)} total &middot; {pages:,} pages_",
      f"",
      f"Generated from `books.json` by `build.py`. Relevance 1-5 (5 = core to research).",
      f"",
      f"| # | Title | Author | Category | Rel | Pages | Status |",
      f"|--:|-------|--------|----------|:---:|------:|--------|"]
for b in books:
    md.append(f"| {b['n']} | {b['title']} | {b['author']} | {b['cat']} | {b['rel']} | {b['pages']} | {b['status']} |")
open(os.path.join(ROOT, "exports", "book_inventory.md"), "w", encoding="utf-8").write("\n".join(md) + "\n")

# ---- exports/Book_Inventory.xlsx ----
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    F = "Arial"
    wb = Workbook(); ws = wb.active; ws.title = "Books"
    heads = ["#", "Title", "Author", "Shelf", "Broad Category", "Relevance (1-5)", "Pages", "Reading Status"]
    ws.append(heads)
    hf = PatternFill("solid", start_color="1F3864"); thin = Side(style="thin", color="D9D9D9"); bd = Border(thin, thin, thin, thin)
    for c in range(1, len(heads) + 1):
        cell = ws.cell(1, c); cell.fill = hf; cell.font = Font(name=F, bold=True, color="FFFFFF", size=11)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True); cell.border = bd
    relf = {5: "C6EFCE", 4: "D6E9C6", 3: "FFEB9C", 2: "FCE4D6", 1: "E7E6E6"}
    relc = {5: "006100", 4: "375623", 3: "7F6000", 2: "843C0C", 1: "595959"}
    stf = {"Finished": ("C6EFCE", "006100"), "Started": ("BDD7EE", "1F4E78"), "Not started": ("F2F2F2", "808080")}
    for i, b in enumerate(books, start=2):
        vals = [b["n"], b["title"], b["author"], b["shelf"], b["cat"], b["rel"], b["pages"], b["status"]]
        for c, v in enumerate(vals, start=1):
            cell = ws.cell(i, c, value=v); cell.font = Font(name=F, size=10); cell.border = bd
            cell.alignment = Alignment(vertical="top", wrap_text=(c in (2, 3, 5)))
        for c in (1, 4, 6, 7, 8):
            ws.cell(i, c).alignment = Alignment(horizontal="center", vertical="top")
        rc = ws.cell(i, 6); rc.fill = PatternFill("solid", start_color=relf[b["rel"]]); rc.font = Font(name=F, size=10, bold=True, color=relc[b["rel"]])
        sc = ws.cell(i, 8); fill, fg = stf[b["status"]]; sc.fill = PatternFill("solid", start_color=fill); sc.font = Font(name=F, size=10, bold=(b["status"] != "Not started"), color=fg)
    for c, w in zip("ABCDEFGH", [4, 40, 30, 7, 24, 14, 9, 14]):
        ws.column_dimensions[c].width = w
    ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:H{len(books)+1}"; ws.row_dimensions[1].height = 40
    wb.save(os.path.join(ROOT, "exports", "Book_Inventory.xlsx"))
    print("wrote index.html, exports/book_inventory.md, exports/Book_Inventory.xlsx")
except ImportError:
    print("wrote index.html, exports/book_inventory.md (openpyxl missing -> skipped xlsx)")
